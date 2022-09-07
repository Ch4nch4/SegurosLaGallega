#como leer un archivo excel

import openpyxl  # libreria buscar en extensiones o 
                 # cdo no la encuentren busquen en internet y
                 # en la INSTALATION les da el pip
                 #  pip install openpyxl

def leerLibroExcel(pathArchivoExcel):
    #leer el archivo
    libro= openpyxl.load_workbook(pathArchivoExcel,data_only=True)
    
    #fijar la hoja
    hoja=libro.active
    
    rango = hoja['A2':'C4'] 
    
    lista_Empleados=[]
    for fila in rango:
        empleado = [celda.value for celda in fila]  #comprension de listas
        lista_Empleados.append(empleado) 
               
    print(lista_Empleados)
    
    #si no conosco el rango???? uso las propiedades (busque en google)
    for fila in hoja.iter_rows(min_col=1,max_col=hoja.max_column):
        print(
            [celda.value for celda in fila]
        )
    
def escribirExcel(pathArchivoExcel):    
    libro=openpyxl.Workbook() 
    #fijar la hoja
    hoja=libro.active
    
   # hoja['A1'] = 'Nombre'
   # hoja['B1'] = 'Apellido'
   # hoja['C1'] = 'Salario'
    hoja.append(['Nombre','Apellido','Salario'])
    
    filas = [      #lista de listitas 
                   # dde cada listita es un empleado
        ['juan','peres',12000],
        ['juana','peres',20000],
        ['pepe','palote',30000]
    ]
    
    for fila in filas:      
        hoja.append(fila)   #cada fila es un empleado

    libro.save(archivo)
##### main
archivo='plantilla.xlsx'
escribirExcel(archivo)
leerLibroExcel(archivo)

#Cómo hacer un Bot de Instagram utilizando Python #1
# CON 4 LINEAS DE CODIGO
#https://www.youtube.com/watch?v=Euhg_XYbDvg

# Stalkea como un pro
# Bot de Instagram #2: Descarga todas las fotos de un usuario

